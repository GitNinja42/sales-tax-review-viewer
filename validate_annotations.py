#!/usr/bin/env python3

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

from run_utils import load_run_manifest, resolve_run_dir
from workbook_utils import choose_review_sheet, extract_transaction_rows, find_default_workbook, find_header_row


VALID_STATUSES = {"Review Needed", "Needs More Context"}
LEGACY_STATUS_MAP = {"Likely Incorrect": "Review Needed"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Validate agent-produced review annotations before writing them to Excel.")
    parser.add_argument("--input-workbook", type=Path, default=None)
    parser.add_argument(
        "--run-dir",
        type=Path,
        default=None,
        help="Run directory created by prepare_review_context.py. Use this to keep validation scoped to one run.",
    )
    parser.add_argument("--annotations", type=Path, default=None)
    parser.add_argument("--control-library", type=Path, default=None)
    return parser.parse_args()


def load_json(path: Path) -> dict[str, object]:
    return json.loads(path.read_text(encoding="utf-8"))


def error(errors: list[str], message: str) -> None:
    errors.append(message)


def main() -> None:
    args = parse_args()
    base_dir = Path(__file__).resolve().parent
    control_library_path = args.control_library or (base_dir / "sales-tax-control-library.json")
    run_dir = None
    manifest = None

    needs_run_dir = args.run_dir is not None or args.input_workbook is None or args.annotations is None
    if needs_run_dir:
        run_dir = resolve_run_dir(base_dir, args.run_dir)
        manifest = load_run_manifest(run_dir)

    manifest_workbook = Path(str(manifest["source_workbook"])) if manifest else find_default_workbook(base_dir)
    input_workbook = args.input_workbook or manifest_workbook
    annotations_path = args.annotations or (run_dir / "annotations.json" if run_dir else None)

    if not annotations_path.exists():
        raise FileNotFoundError(f"Annotation file not found: {annotations_path}")
    if not control_library_path.exists():
        raise FileNotFoundError(f"Control library not found: {control_library_path}")

    annotations = load_json(annotations_path)
    control_library = load_json(control_library_path)
    valid_controls = {item["control_id"] for item in control_library.get("controls", [])}

    wb = load_workbook(input_workbook)
    if annotations.get("source_sheet"):
        source_sheet_name = str(annotations["source_sheet"])
        if source_sheet_name not in wb.sheetnames:
            raise ValueError(f"Source sheet not found in workbook: {source_sheet_name}")
        ws = wb[source_sheet_name]
        header_row = find_header_row(ws)
        if header_row is None:
            raise ValueError(f"Could not find tax-detail headers on source sheet: {source_sheet_name}")
    else:
        ws, header_row, _ = choose_review_sheet(wb)

    headers, transaction_rows = extract_transaction_rows(ws, header_row)
    valid_rows = {int(item["row_num"]) for item in transaction_rows}

    errors: list[str] = []

    required_top_level = {"source_sheet", "output_sheet", "annotation_mode", "row_annotations", "pattern_annotations"}
    missing_top_level = sorted(required_top_level.difference(annotations))
    if missing_top_level:
        error(errors, f"Missing top-level keys: {', '.join(missing_top_level)}")

    if annotations.get("annotation_mode") != "inline_columns":
        error(errors, "annotation_mode must be 'inline_columns'.")

    for index, item in enumerate(annotations.get("row_annotations", []), start=1):
        prefix = f"row_annotations[{index}]"
        if item.get("source_row") not in valid_rows:
            error(errors, f"{prefix}: source_row {item.get('source_row')} is not a valid workbook row.")
        status = item.get("status")
        if status in LEGACY_STATUS_MAP:
            error(
                errors,
                f"{prefix}: status {status!r} is no longer supported for new runs; use {LEGACY_STATUS_MAP[status]!r} instead.",
            )
        elif status not in VALID_STATUSES:
            error(errors, f"{prefix}: invalid status {item.get('status')!r}.")
        refs = item.get("checklist_refs") or []
        if not refs:
            error(errors, f"{prefix}: checklist_refs must contain at least one control ID.")
        for ref in refs:
            if ref not in valid_controls:
                error(errors, f"{prefix}: unknown checklist ref {ref!r}.")
        for column_name in item.get("highlight_columns", []) or []:
            if column_name not in headers:
                error(errors, f"{prefix}: highlight column {column_name!r} does not exist on the source sheet.")

    for index, item in enumerate(annotations.get("pattern_annotations", []), start=1):
        prefix = f"pattern_annotations[{index}]"
        refs = item.get("checklist_refs") or []
        if not refs:
            error(errors, f"{prefix}: checklist_refs must contain at least one control ID.")
        for ref in refs:
            if ref not in valid_controls:
                error(errors, f"{prefix}: unknown checklist ref {ref!r}.")
        row_refs = item.get("row_refs") or []
        if not row_refs:
            error(errors, f"{prefix}: row_refs must contain at least one row number.")
        for row_num in row_refs:
            if row_num not in valid_rows:
                error(errors, f"{prefix}: row_refs contains invalid row number {row_num}.")

    for index, item in enumerate(annotations.get("unmapped_observations", []), start=1):
        prefix = f"unmapped_observations[{index}]"
        row_refs = item.get("row_refs") or []
        for row_num in row_refs:
            if row_num not in valid_rows:
                error(errors, f"{prefix}: row_refs contains invalid row number {row_num}.")

    if errors:
        for message in errors:
            print(f"ERROR: {message}", file=sys.stderr)
        raise SystemExit(1)

    print(f"Annotations validated for sheet '{ws.title}' with {len(annotations.get('row_annotations', []))} row annotations.")


if __name__ == "__main__":
    main()
