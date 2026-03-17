#!/usr/bin/env python3

from __future__ import annotations

import argparse
import json
import warnings
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from run_utils import load_run_manifest, resolve_run_dir, write_run_manifest
from workbook_utils import choose_review_sheet, find_default_workbook, find_header_row

REVIEW_COLUMNS = [
    "Review Status",
    "Issue Summary",
    "Checklist Item(s)",
    "Missing Context",
    "Suggested Next Step",
]

LEGACY_STATUS_MAP = {"Likely Incorrect": "Review Needed"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Apply agent-produced review annotations to an Excel workbook.")
    parser.add_argument("--input-workbook", type=Path, default=None)
    parser.add_argument(
        "--run-dir",
        type=Path,
        default=None,
        help="Run directory created by prepare_review_context.py. Use this to keep workbook output isolated to one run.",
    )
    parser.add_argument("--annotations", type=Path, default=None)
    parser.add_argument("--output-workbook", type=Path, default=None)
    return parser.parse_args()


def load_annotations(path: Path) -> dict[str, object]:
    return json.loads(path.read_text(encoding="utf-8"))


def load_control_titles(path: Path) -> dict[str, str]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    return {item["control_id"]: item["title"] for item in payload.get("controls", [])}


def format_checklist_items(refs: list[str], control_titles: dict[str, str]) -> str:
    ordered_unique_refs = list(dict.fromkeys(refs))
    return "; ".join(control_titles.get(ref, ref) for ref in ordered_unique_refs)


def canonicalize_status(value: object) -> str:
    status = str(value or "").strip()
    return LEGACY_STATUS_MAP.get(status, status or "Review Needed")


def main() -> None:
    warnings.filterwarnings("ignore", message="Title is more than 31 characters.*")
    args = parse_args()
    base_dir = Path(__file__).resolve().parent
    control_library_path = base_dir / "sales-tax-control-library.json"
    run_dir = None
    manifest = None

    needs_run_dir = args.run_dir is not None or args.input_workbook is None or args.annotations is None
    if needs_run_dir:
        run_dir = resolve_run_dir(base_dir, args.run_dir)
        manifest = load_run_manifest(run_dir)

    manifest_workbook = Path(str(manifest["source_workbook"])) if manifest else find_default_workbook(base_dir)
    input_workbook = args.input_workbook or manifest_workbook
    annotations_path = args.annotations or (run_dir / "annotations.json" if run_dir else None)
    if not annotations_path.exists():
        raise FileNotFoundError(f"Annotation file not found: {annotations_path}")
    if not control_library_path.exists():
        raise FileNotFoundError(f"Control library not found: {control_library_path}")

    output_workbook = (
        args.output_workbook
        or (Path(str(manifest["default_reviewed_workbook"])) if manifest else (annotations_path.parent / f"{input_workbook.stem} - Reviewed.xlsx"))
    )
    annotations = load_annotations(annotations_path)
    control_titles = load_control_titles(control_library_path)

    wb = load_workbook(input_workbook)
    if annotations.get("source_sheet"):
        source_sheet = wb[str(annotations["source_sheet"])]
        header_row = find_header_row(source_sheet)
        if header_row is None:
            raise ValueError(f"Could not find the tax-detail header row on source sheet: {source_sheet.title}")
    else:
        source_sheet, header_row, _ = choose_review_sheet(wb)

    review_title = annotations.get("output_sheet") or f"{source_sheet.title} - Reviewed"
    if len(review_title) > 31:
        review_title = "Tax Code - Reviewed"
    if review_title in wb.sheetnames:
        del wb[review_title]
    if "Pattern Flags" in wb.sheetnames:
        del wb["Pattern Flags"]
    review_sheet = wb.copy_worksheet(source_sheet)
    review_sheet.title = review_title

    headers = [review_sheet.cell(header_row, col).value for col in range(1, review_sheet.max_column + 1)]
    header_lookup = {header: idx + 1 for idx, header in enumerate(headers)}
    start_col = review_sheet.max_column + 1

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    status_fill = {
        "Review Needed": PatternFill("solid", fgColor="FFF2CC"),
        "Needs More Context": PatternFill("solid", fgColor="D9E2F3"),
    }
    evidence_fill = {
        "Review Needed": PatternFill("solid", fgColor="FFE599"),
        "Needs More Context": PatternFill("solid", fgColor="9FC5E8"),
    }

    for offset, review_header in enumerate(REVIEW_COLUMNS):
        cell = review_sheet.cell(header_row, start_col + offset, review_header)
        cell.fill = header_fill
        cell.font = header_font

    grouped = defaultdict(list)
    for item in annotations.get("row_annotations", []):
        # Skip rows dismissed by the reviewer.
        if item.get("_dismissed"):
            continue
        normalized = dict(item)
        normalized["status"] = canonicalize_status(item.get("status"))
        grouped[int(item["source_row"])].append(normalized)

    status_rank = {"Needs More Context": 0, "Review Needed": 1}

    for row_num, items in grouped.items():
        status = max((item["status"] for item in items), key=lambda value: status_rank.get(value, -1))
        issue_summary = " | ".join(dict.fromkeys(item["issue_summary"] for item in items))
        all_refs = [ref for item in items for ref in item.get("checklist_refs", [])]
        checklist_refs = format_checklist_items(all_refs, control_titles)
        missing_context = "; ".join(sorted({x for item in items for x in item.get("missing_context", [])}))
        next_step = " | ".join(dict.fromkeys(item.get("suggested_next_step", "") for item in items if item.get("suggested_next_step")))
        highlight_columns = sorted({col for item in items for col in item.get("highlight_columns", [])})

        values = [status, issue_summary, checklist_refs, missing_context, next_step]
        for offset, value in enumerate(values):
            cell = review_sheet.cell(row_num, start_col + offset, value)
            cell.fill = status_fill.get(status, status_fill["Review Needed"])
            if offset == 0:
                cell.font = Font(bold=True)

        for column_name in highlight_columns:
            col_idx = header_lookup.get(column_name)
            if col_idx:
                review_sheet.cell(row_num, col_idx).fill = evidence_fill.get(status, evidence_fill["Review Needed"])

    for index, review_header in enumerate(REVIEW_COLUMNS, start=start_col):
        width = max(len(review_header), 18)
        for row_num in range(header_row + 1, review_sheet.max_row + 1):
            value = review_sheet.cell(row_num, index).value
            if value is not None:
                width = min(max(width, len(str(value)) + 2), 70)
        review_sheet.column_dimensions[get_column_letter(index)].width = width

    pattern_sheet = wb.create_sheet("Pattern Flags")
    pattern_headers = [
        "Pattern ID",
        "Issue Summary",
        "Checklist Item(s)",
        "Row References",
        "Missing Context",
    ]
    for col_idx, header in enumerate(pattern_headers, start=1):
        cell = pattern_sheet.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, item in enumerate(annotations.get("pattern_annotations", []), start=2):
        values = [
            item.get("pattern_id"),
            item.get("issue_summary"),
            format_checklist_items(item.get("checklist_refs", []), control_titles),
            ", ".join(str(x) for x in item.get("row_refs", [])),
            "; ".join(item.get("missing_context", [])),
        ]
        for col_idx, value in enumerate(values, start=1):
            pattern_sheet.cell(row_idx, col_idx, value)

    for col_idx, header in enumerate(pattern_headers, start=1):
        width = max(len(header) + 2, 18)
        for row_idx in range(2, pattern_sheet.max_row + 1):
            value = pattern_sheet.cell(row_idx, col_idx).value
            if value is not None:
                width = min(max(width, len(str(value)) + 2), 80)
        pattern_sheet.column_dimensions[get_column_letter(col_idx)].width = width

    unmapped = annotations.get("unmapped_observations", [])
    if unmapped:
        observation_sheet = wb.create_sheet("Other Observations")
        observation_headers = [
            "Observation ID",
            "Issue Summary",
            "Row References",
            "Why Not Checklist Flag",
            "Suggested Follow-Up",
        ]
        for col_idx, header in enumerate(observation_headers, start=1):
            cell = observation_sheet.cell(1, col_idx, header)
            cell.fill = header_fill
            cell.font = header_font

        for row_idx, item in enumerate(unmapped, start=2):
            values = [
                item.get("observation_id"),
                item.get("issue_summary"),
                ", ".join(str(x) for x in item.get("row_refs", [])),
                item.get("reason_not_checklist_flag"),
                item.get("suggested_follow_up"),
            ]
            for col_idx, value in enumerate(values, start=1):
                observation_sheet.cell(row_idx, col_idx, value)

        for col_idx, header in enumerate(observation_headers, start=1):
            width = max(len(header) + 2, 18)
            for row_idx in range(2, observation_sheet.max_row + 1):
                value = observation_sheet.cell(row_idx, col_idx).value
                if value is not None:
                    width = min(max(width, len(str(value)) + 2), 80)
            observation_sheet.column_dimensions[get_column_letter(col_idx)].width = width

    output_workbook.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_workbook)
    if run_dir is not None:
        write_run_manifest(
            run_dir,
            {
                "annotations_path": str(annotations_path.resolve()),
                "reviewed_workbook": str(output_workbook.resolve()),
                "review_status": "reviewed",
                "review_error": None,
                "review_detail": "The reviewed workbook is ready to download.",
            },
        )
    print(output_workbook)


if __name__ == "__main__":
    main()
